import pandas as pd
import logging
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher
import numpy as np

def setup_logging():
    """Configura il sistema di logging"""
    log_dir = Path('logs')
    log_dir.mkdir(exist_ok=True)
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_dir / f'merge_{datetime.now().strftime("%Y%m%d")}.log'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

def carica_excel_originale(path_excel: str) -> pd.DataFrame:
    """Carica l'Excel originale preservando i formati"""
    if not Path(path_excel).exists():
        raise FileNotFoundError(f"File Excel non trovato: {path_excel}")
    
    # Carica il file Excel preservando i formati
    wb = openpyxl.load_workbook(path_excel)
    ws = wb.active
    
    # Estrai i dati mantenendo i formati
    dati = []
    headers = [cell.value for cell in ws[1]]
    
    for row in ws.iter_rows(min_row=2):
        riga = {}
        for cell, header in zip(row, headers):
            # Preserva il formato della cella
            riga[header] = {
                'valore': cell.value,
                'formato': {
                    'font': cell.font,
                    'fill': cell.fill,
                    'alignment': cell.alignment,
                    'border': cell.border
                }
            }
        dati.append(riga)
    
    return pd.DataFrame(dati), wb, ws

def carica_dati_scraped():
    """Carica l'ultimo dataset estratto"""
    final_dir = Path('data/final')
    if not final_dir.exists():
        raise FileNotFoundError("Directory data/final non trovata")
    
    # Trova l'ultimo file CSV
    csv_files = list(final_dir.glob('missioni_internazionali_*.csv'))
    if not csv_files:
        raise FileNotFoundError("Nessun file CSV trovato in data/final")
    
    latest_file = max(csv_files, key=lambda x: x.stat().st_mtime)
    return pd.read_csv(latest_file)

def mappa_colonne(df_scraped: pd.DataFrame, df_originale: pd.DataFrame) -> dict:
    """Mappa le colonne tra i due DataFrame"""
    # Mappa standard delle colonne
    mappa = {
        'nome_missione': 'Nome Missione',
        'paese': 'Paese',
        'data_inizio': 'Data Inizio',
        'data_fine': 'Data Fine',
        'personale_totale': 'Personale',
        'costo_totale': 'Costo_€',
        'tipo_missione': 'Tipo Missione',
        'mandato': 'Mandato',
        'note': 'Note',
        'link_documento': 'Link Documento'
    }
    
    # Verifica che tutte le colonne mappate esistano
    for col_scraped, col_originale in mappa.items():
        if col_scraped not in df_scraped.columns:
            logging.warning(f"Colonna {col_scraped} non trovata nei dati scraped")
        if col_originale not in df_originale.columns:
            logging.warning(f"Colonna {col_originale} non trovata nell'Excel originale")
    
    return mappa

def calcola_similarita(str1: str, str2: str) -> float:
    """Calcola la similarità tra due stringhe"""
    if pd.isna(str1) or pd.isna(str2):
        return 0.0
    return SequenceMatcher(None, str(str1).lower(), str(str2).lower()).ratio()

def trova_duplicati(df_scraped: pd.DataFrame, df_originale: pd.DataFrame, mappa: dict) -> pd.DataFrame:
    """Trova e gestisce i duplicati tra i due DataFrame"""
    # Crea una colonna di similarità per nome missione e paese
    df_scraped['similarita'] = df_scraped.apply(
        lambda row: df_originale.apply(
            lambda orig_row: max(
                calcola_similarita(row['nome_missione'], orig_row[mappa['nome_missione']]),
                calcola_similarita(row['paese'], orig_row[mappa['paese']])
            ),
            axis=1
        ).max(),
        axis=1
    )
    
    # Identifica i duplicati (similarità > 0.8)
    duplicati = df_scraped[df_scraped['similarita'] > 0.8].copy()
    
    # Per ogni duplicato, trova la riga corrispondente nell'originale
    duplicati['riga_originale'] = duplicati.apply(
        lambda row: df_originale.index[
            df_originale.apply(
                lambda orig_row: max(
                    calcola_similarita(row['nome_missione'], orig_row[mappa['nome_missione']]),
                    calcola_similarita(row['paese'], orig_row[mappa['paese']])
                ),
                axis=1
            ).argmax()
        ],
        axis=1
    )
    
    return duplicati

def unisci_dati(df_scraped: pd.DataFrame, df_originale: pd.DataFrame, wb: openpyxl.Workbook, ws: openpyxl.worksheet.worksheet.Worksheet):
    """Unisce i dati preservando i formati dell'Excel originale"""
    # Mappa le colonne
    mappa = mappa_colonne(df_scraped, df_originale)
    
    # Trova i duplicati
    duplicati = trova_duplicati(df_scraped, df_originale, mappa)
    
    # Aggiorna i valori mantenendo i formati
    for idx, row in df_scraped.iterrows():
        # Se è un duplicato, usa la riga originale
        if idx in duplicati.index:
            row_idx = duplicati.loc[idx, 'riga_originale'] + 2  # +2 perché l'Excel ha l'header e è 1-based
        else:
            # Trova la riga corrispondente nell'Excel
            mask = df_originale['Nome Missione'] == row['nome_missione']
            if mask.any():
                row_idx = mask.idxmax() + 2
            else:
                # Se non trova corrispondenza, aggiungi una nuova riga
                row_idx = ws.max_row + 1
        
        # Aggiorna i valori mantenendo i formati
        for col_scraped, col_originale in mappa.items():
            if col_scraped in row and col_originale in df_originale.columns:
                cell = ws[f"{get_column_letter(df_originale.columns.get_loc(col_originale) + 1)}{row_idx}"]
                
                # Aggiorna il valore mantenendo il formato
                cell.value = row[col_scraped]
                
                # Se è un duplicato, evidenzia la cella
                if idx in duplicati.index:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    return wb

def salva_excel_aggiornato(wb: openpyxl.Workbook, path_output: str):
    """Salva il file Excel aggiornato"""
    wb.save(path_output)
    logging.info(f"File Excel aggiornato salvato in: {path_output}")

def main():
    logger = setup_logging()
    logger.info("Avvio processo di merge con Excel")
    
    try:
        # Carica i dati
        logger.info("Caricamento dati scraped")
        df_scraped = carica_dati_scraped()
        
        # Path del file Excel master
        path_excel = "data/final/Matrice dati 1AGG.xlsx"
        logger.info(f"Caricamento file Excel master: {path_excel}")
        df_originale, wb, ws = carica_excel_originale(path_excel)
        
        # Unisci i dati
        logger.info("Unione dei dati")
        wb_aggiornato = unisci_dati(df_scraped, df_originale, wb, ws)
        
        # Salva il file aggiornato
        path_output = f"data/final/Matrice dati 1AGG_aggiornato_{datetime.now().strftime('%Y%m%d')}.xlsx"
        salva_excel_aggiornato(wb_aggiornato, path_output)
        
        logger.info("Processo di merge completato con successo")
        
    except Exception as e:
        logger.error(f"Errore durante il merge: {str(e)}")
        raise

if __name__ == "__main__":
    main() 